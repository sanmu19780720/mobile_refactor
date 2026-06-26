from __future__ import annotations

import io
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from app.services.excel import SUPPLIER_NAME, export_mode_label

TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "templates"
BATCH_TEMPLATE = TEMPLATE_DIR / "batch_import.xlsx"
COATS_TEMPLATE = TEMPLATE_DIR / "coats_order.xlsx"


def _f(line: Dict[str, Any], key: str, default: float = 0.0) -> float:
    val = line.get(key, default)
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default


def final_price(price: float, discount: float, jiajia: float) -> float:
    """折后价 = 原价 ×(100+折扣)/100 + 客户加价，保留 2 位四舍五入。

    与前端 finalPrice 保持一致；佣金比例为内部用、不计入。
    """
    base = price * (100 + discount) / 100 + jiajia
    return float(Decimal(str(base)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _reset_keep_header(ws, header_rows: int = 1) -> Dict[int, str]:
    """清掉模版里的示例数据行，仅保留表头；返回首个数据行的各列 number_format 以便复用。"""
    sample_row = header_rows + 1
    formats: Dict[int, str] = {}
    if ws.max_row >= sample_row:
        for col in range(1, ws.max_column + 1):
            formats[col] = ws.cell(row=sample_row, column=col).number_format
        ws.delete_rows(sample_row, ws.max_row - header_rows)
    return formats


def _write_row(ws, row_idx: int, values: List[Any], formats: Dict[int, str]) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        fmt = formats.get(col)
        if fmt:
            cell.number_format = fmt


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_batch_import_xlsx(lines: List[Dict[str, Any]]) -> bytes:
    """批量导入表：直接套用模版表头/格式，逐行写入临时报价数据。

    列：序号 / 供应商名称 / 客户名称 / 客户款号 / 客户订单号码 / ITEMCODE /
        颜色代号 / 数量 / 折扣 / 佣金方式 / 佣金比例 / 客户加价 / 代管卖方名称 / 代管折扣
    """
    wb = load_workbook(BATCH_TEMPLATE)
    ws = wb.active
    formats = _reset_keep_header(ws)

    row_idx = 2
    for idx, ln in enumerate(lines, start=1):
        yj_p = _f(ln, "cust_yongjin_p")
        zk_jj = _f(ln, "cust_zhekou_jiajia")
        serial = ln.get("serial_no") or idx
        _write_row(
            ws,
            row_idx,
            [
                serial,
                SUPPLIER_NAME,
                ln.get("cust_name", "") or "",
                ln.get("cust_kuanhao", "") or "",
                ln.get("cust_po", "") or "",
                ln.get("itemcode", "") or "",
                ln.get("color", "") or "",
                int(ln.get("qty", 0) or 0),
                _f(ln, "discount"),
                export_mode_label(yj_p, zk_jj),
                yj_p,
                zk_jj,
                "",
                "",
            ],
            formats,
        )
        row_idx += 1

    return _save(wb)


def build_coats_xlsx(
    lines: List[Dict[str, Any]],
    ship_to: str,
    buyer: str,
    po_prefix: str,
    start_seq: int,
    required_date: date,
) -> bytes:
    """高士下单表：套用模版，PO No. 逐行递增，收件地址/买家手填，送货日期=下单日期+5天。

    列：PO No.（客户单号）/ Ship to Party（收件地址）/ Buyer（买家）/
        Required Date（要求送货日期）/ Article（货号）/ Shade（颜色编号）/ Qty（数量）/
        Vendor Code / Brand Order Date / Brand Priority
    """
    wb = load_workbook(COATS_TEMPLATE)
    ws = wb.active
    formats = _reset_keep_header(ws)

    row_idx = 2
    for i, ln in enumerate(lines):
        po_no = int(f"{po_prefix}{(start_seq + i):04d}")
        _write_row(
            ws,
            row_idx,
            [
                po_no,
                ship_to,
                buyer,
                required_date,
                ln.get("itemcode", "") or "",
                ln.get("color", "") or "",
                int(ln.get("qty", 0) or 0),
                "",
                "",
                "",
            ],
            formats,
        )
        row_idx += 1

    return _save(wb)
